#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Advanced Excel Generator with Image and Border Support
画像挿入と罫線をサポートする高度なExcel生成スクリプト
"""

import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

def create_inspection_report(data, output_path):
    """
    点検帳票を生成
    
    Args:
        data: 点検データ（JSON形式）
        output_path: 出力ファイルパス
    """
    
    # ワークブック作成
    wb = Workbook()
    ws = wb.active
    ws.title = '油圧ｼｮﾍﾞﾙ'
    
    # 基本情報の取得
    machine_type = data.get('machine_type', '油圧ショベル')
    machine_model = data.get('machine_model', '')
    machine_unit = data.get('machine_unit', '')
    site_name = data.get('site_name', '')
    company_name = data.get('company_name', '')
    responsible_person = data.get('responsible_person', '')
    prime_contractor_inspector = data.get('prime_contractor_inspector', '') # 元請点検責任者
    month = data.get('month', 1)
    year = data.get('year', 2025)
    records = data.get('records', [])
    items = data.get('items', [])
    
    # 型式部分のみ抽出（括弧内の文字）
    # 全角括弧と半角括弧の両方に対応
    # 例: 「油圧ショベル（PC200）」→「PC200」
    # 例: 「油圧ショベル(PC200)」→「PC200」
    model_spec = ''
    if '（' in machine_model and '）' in machine_model:
        start_idx = machine_model.index('（') + 1
        end_idx = machine_model.index('）')
        model_spec = machine_model[start_idx:end_idx]
    elif '(' in machine_model and ')' in machine_model:
        start_idx = machine_model.index('(') + 1
        end_idx = machine_model.index(')')
        model_spec = machine_model[start_idx:end_idx]
    
    # 重機名部分のみ抽出（括弧の前まで）
    # 全角括弧と半角括弧の両方に対応
    # 例: 「油圧ショベル（PC200）」→「油圧ショベル」
    # 例: 「油圧ショベル(PC200)」→「油圧ショベル」
    machine_name = machine_type
    if '（' in machine_name:
        machine_name = machine_name[:machine_name.index('（')]
    elif '(' in machine_name:
        machine_name = machine_name[:machine_name.index('(')]
    
    # ============================================================
    # スタイル定義
    # ============================================================
    
    # フォント
    font_hgmincho_18 = Font(name='HG明朝E', size=18)
    font_hgmincho_14 = Font(name='HG明朝E', size=14)
    font_hgmincho_14_bold = Font(name='HG明朝E', size=14, bold=True)
    font_hgmincho_16_bold_underline = Font(name='HG明朝E', size=16, bold=True, underline='single')
    font_hgmincho_22_bold = Font(name='HG明朝E', size=22, bold=True)
    font_hgmincho_26_bold_italic = Font(name='HG明朝E', size=26, bold=True, italic=True)
    font_hgmincho_16 = Font(name='HG明朝E', size=16)
    font_hgmincho_12 = Font(name='HG明朝E', size=12)
    font_hgmincho_11 = Font(name='HG明朝E', size=11, bold=True)
    font_hgmincho_10 = Font(name='HG明朝E', size=10)
    font_hgmincho_9 = Font(name='HG明朝E', size=9)
    
    # 配置
    align_left_center = Alignment(horizontal='left', vertical='center')
    align_center_center = Alignment(horizontal='center', vertical='center')
    align_left_bottom = Alignment(horizontal='left', vertical='bottom')
    align_center_center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 背景色
    fill_gray = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    fill_green = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    fill_red = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
    
    # 罫線
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    top_border = Border(top=Side(style='thin'))
    left_border = Border(left=Side(style='thin'))
    bottom_border = Border(bottom=Side(style='thin'))
    right_border = Border(right=Side(style='thin'))
    
    # ============================================================
    # 列幅設定
    # ============================================================
    
    # A列: 36px ≈ 5文字
    ws.column_dimensions['A'].width = 5.0
    
    # B～AK列: 24px ≈ 3.3文字
    for col_idx in range(2, 38):
        ws.column_dimensions[get_column_letter(col_idx)].width = 3.3
    
    # AL列: 48px ≈ 6.7文字
    ws.column_dimensions['AL'].width = 6.7
    
    # AM列以降: 32px ≈ 4.5文字
    for col_idx in range(39, 71):
        ws.column_dimensions[get_column_letter(col_idx)].width = 4.5
    
    # ============================================================
    # 行高設定
    # ============================================================
    
    # 行1～4: 24px
    for row in range(1, 5):
        ws.row_dimensions[row].height = 24
    
    # 行5: 43px
    ws.row_dimensions[5].height = 43
    
    # 行6: 18px
    ws.row_dimensions[6].height = 18
    
    # 行7: 31px
    ws.row_dimensions[7].height = 31
    
    # 行8: 9px
    ws.row_dimensions[8].height = 9
    
    # 行9～26: 32px
    for row in range(9, 27):
        ws.row_dimensions[row].height = 32
    
    # 行27: 72px
    ws.row_dimensions[27].height = 72
    
    # 行28～31: 37px
    for row in range(28, 32):
        ws.row_dimensions[row].height = 37
    
    # ============================================================
    # 行1: 工事名
    # ============================================================
    
    ws['A1'] = '工事名'
    ws['A1'].font = font_hgmincho_18
    ws['A1'].alignment = align_left_center
    
    # D1:E1を結合して「：」
    ws.merge_cells('D1:E1')
    ws['D1'] = '：'
    ws['D1'].font = font_hgmincho_14
    ws['D1'].alignment = align_center_center
    
    # F1に工事名
    if site_name:
        ws['F1'] = site_name
        ws['F1'].font = font_hgmincho_18
        ws['F1'].alignment = align_left_center
    
    # ============================================================
    # 行3: 法的要求事項とヘッダー情報
    # ============================================================
    
    # A3セルの関係法令を重機種類ごとに設定
    if '油圧ショベル' in machine_type or '油圧ｼｮﾍﾞﾙ' in machine_type:
        ws['A3'] = '　【ｸﾚｰﾝ則第７８条】'
    elif 'ハンドガイド式' in machine_type:
        ws['A3'] = '　【労働安全衛生法第２０条】'
    else:
        # ブルドーザ、不整地運搬車、コンバインドローラー、振動ローラー
        ws['A3'] = '　【安衛則第１７０条】'
    ws['A3'].font = font_hgmincho_14
    ws['A3'].alignment = align_left_center
    
    ws['J3'] = '・★は法的要求事項'
    ws['J3'].font = font_hgmincho_14
    ws['J3'].alignment = align_left_center
    
    # AM3:AW3 → 所有会社名
    ws.merge_cells('AM3:AW3')
    ws['AM3'] = '所有会社名'
    ws['AM3'].font = font_hgmincho_11
    ws['AM3'].alignment = align_center_center
    
    # AX3:BD3 → 取扱責任者（点検者）
    ws.merge_cells('AX3:BD3')
    ws['AX3'] = '取扱責任者（点検者）'
    ws['AX3'].font = font_hgmincho_11
    ws['AX3'].alignment = align_center_center
    
    # BE3:BH3 → 型式
    ws.merge_cells('BE3:BH3')
    ws['BE3'] = '型式'
    ws['BE3'].font = font_hgmincho_11
    ws['BE3'].alignment = align_center_center
    
    # BI3:BL3 → 機械番号
    ws.merge_cells('BI3:BL3')
    ws['BI3'] = '機械番号'
    ws['BI3'].font = font_hgmincho_11
    ws['BI3'].alignment = align_center_center
    
    # BN3:BQ3 → 作業所長確認
    ws.merge_cells('BN3:BQ3')
    ws['BN3'] = '作業所長確認'
    ws['BN3'].font = font_hgmincho_11
    ws['BN3'].alignment = align_center_center
    
    # ============================================================
    # 行4: 法的要求事項と型式・号機
    # ============================================================
    
    # A4セルの関係法令（油圧ショベルの場合のみ記入）
    if '油圧ショベル' in machine_type or '油圧ｼｮﾍﾞﾙ' in machine_type:
        ws['A4'] = '　【安衛則第１７０条】'
        ws['A4'].font = font_hgmincho_14
        ws['A4'].alignment = align_left_center
    
    ws['J4'] = '・その他は点検すべき事項とみなした箇所'
    ws['J4'].font = font_hgmincho_14
    ws['J4'].alignment = align_left_center
    
    # AM4:AW5を結合（所有会社名記入欄）
    ws.merge_cells('AM4:AW5')
    ws['AM4'] = company_name
    ws['AM4'].font = font_hgmincho_16
    ws['AM4'].alignment = align_center_center
    
    # AX4:BD5を結合（取扱責任者（点検者）名記入欄）
    ws.merge_cells('AX4:BD5')
    ws['AX4'] = responsible_person
    ws['AX4'].font = font_hgmincho_16
    ws['AX4'].alignment = align_center_center
    
    # BE4:BH5を結合して型式を自動入力（中央配置）
    ws.merge_cells('BE4:BH5')
    ws['BE4'] = machine_model  # machine_model全体を使用（括弧内のみではない）
    ws['BE4'].font = font_hgmincho_16
    ws['BE4'].alignment = align_center_center
    
    # BI4:BL5を結合して号機番号を自動入力
    ws.merge_cells('BI4:BL5')
    ws['BI4'] = machine_unit
    ws['BI4'].font = font_hgmincho_16
    ws['BI4'].alignment = align_center_center
    
    # BN4:BQ5の結合は後で一括実行（行655）
    
    # ============================================================
    # 行5: タイトル（重機名のみ、型式は除外）
    # ============================================================
    
    ws['A5'] = f'{month}月度　{machine_name}　作業開始前点検表'
    ws['A5'].font = font_hgmincho_26_bold_italic
    ws['A5'].alignment = align_left_bottom
    
    # ============================================================
    # 行7: 注意書き
    # ============================================================
    
    ws['A7'] = '※点検時、作業時問わず異常を認めたときは、元請点検責任者に報告及び速やかに補修その他必要な措置を取ること'
    ws['A7'].font = font_hgmincho_16_bold_underline
    ws['A7'].alignment = align_left_bottom
    
    # ============================================================
    # 行9: ヘッダー行
    # ============================================================
    
    # A9:Q9 → 点検項目
    ws.merge_cells('A9:Q9')
    ws['A9'] = '点検項目'
    ws['A9'].font = font_hgmincho_14_bold
    ws['A9'].fill = fill_gray
    ws['A9'].alignment = align_center_center
    
    # R9:AL9 → 点検ポイント
    ws.merge_cells('R9:AL9')
    ws['R9'] = '点検ポイント'
    ws['R9'].font = font_hgmincho_14_bold
    ws['R9'].fill = fill_gray
    ws['R9'].alignment = align_center_center
    
    # AM9～BQ9に1～31日（11pt）
    for day in range(1, 32):
        col_idx = 38 + day  # AM列は39列目（38+1）
        if col_idx <= 69:
            cell = ws.cell(row=9, column=col_idx)
            cell.value = str(day)
            cell.font = font_hgmincho_11
            cell.fill = fill_gray
            cell.alignment = align_center_center
    
    # A9～BQ9の上部に罫線
    for col_idx in range(1, 70):
        cell = ws.cell(row=9, column=col_idx)
        cell.border = top_border
    
    # ============================================================
    # 行10～23: 点検項目とデータ
    # ============================================================
    
    for i, item in enumerate(items[:14]):
        row = 10 + i
        
        # A列: ★マーク
        if item.get('is_required', False):
            ws.cell(row=row, column=1).value = '★'
            ws.cell(row=row, column=1).font = font_hgmincho_14
            ws.cell(row=row, column=1).alignment = align_center_center
        
        # B列: 項目名
        ws.cell(row=row, column=2).value = item.get('name', '')
        ws.cell(row=row, column=2).font = font_hgmincho_14
        ws.cell(row=row, column=2).alignment = align_left_center
        
        # R列: 点検ポイント
        ws.cell(row=row, column=18).value = item.get('check_point', '')
        ws.cell(row=row, column=18).font = font_hgmincho_14
        ws.cell(row=row, column=18).alignment = align_left_center
        
        # AM列～: ⚪×データ
        item_code = item.get('code', '')
        for record in records:
            day = int(record.get('day', 0))
            results = record.get('results', {})
            
            if item_code in results:
                col_idx = 38 + day
                if col_idx <= 69:
                    cell = ws.cell(row=row, column=col_idx)
                    result = results[item_code]
                    
                    # is_goodの判定（True、true、1、'1'、'true'を良好として扱う）
                    is_good_value = result.get('is_good')
                    if is_good_value == True or is_good_value == 'true' or is_good_value == 1 or is_good_value == '1':
                        cell.value = '⚪'
                        cell.fill = fill_green
                    elif is_good_value == False or is_good_value == 'false' or is_good_value == 0 or is_good_value == '0':
                        cell.value = '×'
                        cell.fill = fill_red
                    
                    cell.font = font_hgmincho_10
                    cell.alignment = align_center_center
    
    # ============================================================
    # 行24～26: 説明と点検者
    # ============================================================
    
    ws['A24'] = '１．点検時'
    ws['A24'].font = font_hgmincho_14
    ws['A24'].alignment = align_left_center
    
    ws['I24'] = '良好…○　要調整、修理…×（使用禁止）　・該当なし…－'
    ws['I24'].font = font_hgmincho_14
    ws['I24'].alignment = align_left_center
    
    ws['B25'] = 'チェック記号'
    ws['B25'].font = font_hgmincho_14
    ws['B25'].alignment = align_left_center
    
    ws['I25'] = '調整または補修したとき…×を○で囲む'
    ws['I25'].font = font_hgmincho_14
    ws['I25'].alignment = align_left_center
    
    ws['A26'] = '２．元請点検責任者は、毎月上旬・中旬・下旬毎に１回点検状況を確認すること。'
    ws['A26'].font = font_hgmincho_14
    ws['A26'].alignment = align_left_center
    
    # AL24:AL26 → 点検者
    ws.merge_cells('AL24:AL26')
    ws['AL24'] = '点\n検\n者'
    ws['AL24'].font = font_hgmincho_12
    ws['AL24'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # AM24～BQ26: 点検者名
    for record in records:
        day = int(record.get('day', 0))
        inspector = record.get('inspector_name', '')
        col_idx = 38 + day
        
        if col_idx <= 69:
            # 24～26行を結合
            start_cell = f'{get_column_letter(col_idx)}24'
            end_cell = f'{get_column_letter(col_idx)}26'
            ws.merge_cells(f'{start_cell}:{end_cell}')
            
            cell = ws[start_cell]
            # 点検者名を縦書きにするため、各文字を改行で区切る
            vertical_text = '\n'.join(list(inspector))
            cell.value = vertical_text
            cell.font = font_hgmincho_10
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # ============================================================
    # 行27: 元請点検責任者確認欄
    # ============================================================
    
    # AK27:AL27
    ws.merge_cells('AK27:AL27')
    ws['AK27'] = '元請点検\n責任者\n確認欄'
    ws['AK27'].font = font_hgmincho_10
    ws['AK27'].alignment = align_center_center_wrap
    
    # AM27:AT27 - 元請点検責任者を記入（16pt、中央配置）
    ws.merge_cells('AM27:AT27')
    ws['AM27'] = prime_contractor_inspector
    ws['AM27'].font = font_hgmincho_16
    ws['AM27'].alignment = align_center_center
    
    # AU27:AV27を結合
    ws.merge_cells('AU27:AV27')
    
    # AW27:BD27 - 元請点検責任者を記入（16pt、中央配置）
    ws.merge_cells('AW27:BD27')
    ws['AW27'] = prime_contractor_inspector
    ws['AW27'].font = font_hgmincho_16
    ws['AW27'].alignment = align_center_center
    
    # BE27:BF27を結合
    ws.merge_cells('BE27:BF27')
    
    # BG27:BO27 - 元請点検責任者を記入（16pt、中央配置）
    ws.merge_cells('BG27:BO27')
    ws['BG27'] = prime_contractor_inspector
    ws['BG27'].font = font_hgmincho_16
    ws['BG27'].alignment = align_center_center
    
    # BP27:BQ27を結合
    ws.merge_cells('BP27:BQ27')
    
    # ============================================================
    # 行28: 補修関連ヘッダー（11pt）
    # ============================================================
    
    ws.merge_cells('AK28:BE28')
    ws['AK28'] = '補修内容'
    ws['AK28'].font = font_hgmincho_11
    ws['AK28'].alignment = align_center_center
    
    ws.merge_cells('BF28:BH28')
    ws['BF28'] = '補修日'
    ws['BF28'].font = font_hgmincho_11
    ws['BF28'].alignment = align_center_center
    
    ws.merge_cells('BI28:BK28')
    ws['BI28'] = '補修者'
    ws['BI28'].font = font_hgmincho_11
    ws['BI28'].alignment = align_center_center
    
    ws.merge_cells('BL28:BN28')
    ws['BL28'] = '元請点検\n責任者'
    ws['BL28'].font = font_hgmincho_11
    ws['BL28'].alignment = align_center_center_wrap
    
    ws.merge_cells('BO28:BQ28')
    ws['BO28'] = '作業所長'
    ws['BO28'].font = font_hgmincho_11
    ws['BO28'].alignment = align_center_center
    
    # ============================================================
    # 行29: 補修関連データ行（セル結合）
    # ============================================================
    
    ws.merge_cells('AK29:BE29')
    ws.merge_cells('BF29:BH29')
    ws.merge_cells('BI29:BK29')
    ws.merge_cells('BL29:BN29')
    ws.merge_cells('BO29:BQ29')
    
    # ============================================================
    # 行30～31: 補修関連データ行（セル結合）
    # ============================================================
    
    ws.merge_cells('AK30:BE30')
    ws.merge_cells('BF30:BH30')
    ws.merge_cells('BI30:BK30')
    ws.merge_cells('BL30:BN30')  # 行30のBL～BN列を結合
    ws.merge_cells('BO30:BQ30')
    
    ws.merge_cells('AK31:BE31')
    ws.merge_cells('BF31:BH31')
    ws.merge_cells('BI31:BK31')
    ws.merge_cells('BL31:BN31')  # 行31のBL～BN列を結合
    ws.merge_cells('BO31:BQ31')
    
    # ============================================================
    # 行27～31: 重機画像エリア（A27:AJ31を結合）
    # ============================================================
    
    # A27:AJ31をマージし、「※重機画像添付※」と表示
    ws.merge_cells('A27:AJ31')
    ws['A27'] = '※重機画像添付※'
    ws['A27'].font = Font(name='HG明朝E', size=18, bold=True)
    ws['A27'].alignment = align_center_center
    
    # ============================================================
    # 罫線の設定（A9～BQ31の外枠）
    # ============================================================
    
    # A～BQ列の行10～24の各セル上部に罫線を追加
    for row in range(10, 25):
        for col_idx in range(1, 70):  # A列(1)～BQ列(69)
            cell = ws.cell(row=row, column=col_idx)
            current_border = cell.border if cell.border else Border()
            cell.border = Border(
                top=Side(style='thin'),
                left=current_border.left,
                right=current_border.right,
                bottom=current_border.bottom
            )
    
    # A9～A31の左側に罫線
    for row in range(9, 32):
        ws.cell(row=row, column=1).border = Border(
            left=Side(style='thin'),
            top=ws.cell(row=row, column=1).border.top if ws.cell(row=row, column=1).border else None,
            bottom=ws.cell(row=row, column=1).border.bottom if ws.cell(row=row, column=1).border else None,
            right=ws.cell(row=row, column=1).border.right if ws.cell(row=row, column=1).border else None
        )
    
    # A31～BQ31の下部に罫線
    for col_idx in range(1, 70):
        ws.cell(row=31, column=col_idx).border = Border(
            bottom=Side(style='thin'),
            left=ws.cell(row=31, column=col_idx).border.left if ws.cell(row=31, column=col_idx).border else None,
            top=ws.cell(row=31, column=col_idx).border.top if ws.cell(row=31, column=col_idx).border else None,
            right=ws.cell(row=31, column=col_idx).border.right if ws.cell(row=31, column=col_idx).border else None
        )
    
    # BQ9～BQ31の右側に罫線
    for row in range(9, 32):
        ws.cell(row=row, column=69).border = Border(
            right=Side(style='thin'),
            left=ws.cell(row=row, column=69).border.left if ws.cell(row=row, column=69).border else None,
            top=ws.cell(row=row, column=69).border.top if ws.cell(row=row, column=69).border else None,
            bottom=ws.cell(row=row, column=69).border.bottom if ws.cell(row=row, column=69).border else None
        )
    
    # ============================================================
    # 追加罫線設定
    # ============================================================
    
    # 1. 行10～23のA列、Q列、AL列の右側に罫線
    for row in range(10, 24):
        # A列(1)の右側
        cell = ws.cell(row=row, column=1)
        cell.border = Border(
            left=cell.border.left if cell.border else None,
            top=cell.border.top if cell.border else None,
            bottom=cell.border.bottom if cell.border else None,
            right=Side(style='thin')
        )
        # Q列(17)の右側
        cell = ws.cell(row=row, column=17)
        cell.border = Border(
            left=cell.border.left if cell.border else None,
            top=cell.border.top if cell.border else None,
            bottom=cell.border.bottom if cell.border else None,
            right=Side(style='thin')
        )
        # AL列(38)の右側
        cell = ws.cell(row=row, column=38)
        cell.border = Border(
            left=cell.border.left if cell.border else None,
            top=cell.border.top if cell.border else None,
            bottom=cell.border.bottom if cell.border else None,
            right=Side(style='thin')
        )
    
    # 2. AM～BL列(38～64)の行2、3、5の下部に罫線
    for row in [2, 3, 5]:
        for col_idx in range(38, 65):  # AM(38)～BL(64)
            cell = ws.cell(row=row, column=col_idx)
            cell.border = Border(
                left=cell.border.left if cell.border else None,
                top=cell.border.top if cell.border else None,
                right=cell.border.right if cell.border else None,
                bottom=Side(style='thin')
            )
    
    # 2-0. AL2セルの下部罫線を削除
    cell = ws.cell(row=2, column=38)
    cell.border = Border(
        left=cell.border.left if cell.border else None,
        top=cell.border.top if cell.border else None,
        right=cell.border.right if cell.border else None
    )
    
    # 3. 行3～5のAL、AW、BD、BH、BL、BM、BQ列の右側に罫線
    # AL=38, AW=49, BD=56, BH=60, BL=64, BM=65, BQ=69
    border_cols = [38, 49, 56, 60, 64, 65, 69]
    for row in range(3, 6):
        for col_idx in border_cols:
            cell = ws.cell(row=row, column=col_idx)
            cell.border = Border(
                left=cell.border.left if cell.border else None,
                top=cell.border.top if cell.border else None,
                bottom=cell.border.bottom if cell.border else None,
                right=Side(style='thin')
            )
    
    # 4. A～AK列(1～37)の行25、26の下部に罫線
    for row in [25, 26]:
        for col_idx in range(1, 38):  # A(1)～AK(37)
            cell = ws.cell(row=row, column=col_idx)
            cell.border = Border(
                left=cell.border.left if cell.border else None,
                top=cell.border.top if cell.border else None,
                right=cell.border.right if cell.border else None,
                bottom=Side(style='thin')
            )
    
    # 5. AK～BQ列(37～69)の行27、28、29、30の下部に罫線
    for row in [27, 28, 29, 30]:
        for col_idx in range(37, 70):  # AK(37)～BQ(69)
            cell = ws.cell(row=row, column=col_idx)
            cell.border = Border(
                left=cell.border.left if cell.border else None,
                top=cell.border.top if cell.border else None,
                right=cell.border.right if cell.border else None,
                bottom=Side(style='thin')
            )
    
    # 6. 行27のAL、AV、BF列の右側に罫線
    # AL=38, AV=48, BF=58
    for col_idx in [38, 48, 58]:
        cell = ws.cell(row=27, column=col_idx)
        cell.border = Border(
            left=cell.border.left if cell.border else None,
            top=cell.border.top if cell.border else None,
            bottom=cell.border.bottom if cell.border else None,
            right=Side(style='thin')
        )
    
    # 7. 行28～31のBE、BH、BK、BN列の右側に罫線
    # BE=57, BH=60, BK=63, BN=66
    for row in range(28, 32):
        for col_idx in [57, 60, 63, 66]:
            cell = ws.cell(row=row, column=col_idx)
            cell.border = Border(
                left=cell.border.left if cell.border else None,
                top=cell.border.top if cell.border else None,
                bottom=cell.border.bottom if cell.border else None,
                right=Side(style='thin')
            )
    
    # ============================================================
    # 追加修正: 罫線の追加・削除
    # ============================================================
    
    # 1. 行3～5のAL列（38列）: 右側以外の罫線を削除
    for row in range(3, 6):
        cell = ws.cell(row=row, column=38)
        cell.border = Border(
            right=Side(style='thin')
        )
    
    # 1-1. AL列の行3上部の罫線を削除
    cell = ws.cell(row=3, column=38)
    cell.border = Border(
        right=Side(style='thin')
    )
    
    # 1-2. A列～Y列の行3、行4に罫線を追加
    # A列の行3、4の左側罫線
    for row in [3, 4]:
        cell = ws.cell(row=row, column=1)
        current_border = cell.border
        cell.border = Border(
            left=Side(style='thin'),
            top=current_border.top if current_border else None,
            right=current_border.right if current_border else None,
            bottom=current_border.bottom if current_border else None
        )
    
    # A～Y列の行3上部罫線
    for col_idx in range(1, 26):  # A(1)～Y(25)
        cell = ws.cell(row=3, column=col_idx)
        current_border = cell.border
        cell.border = Border(
            left=current_border.left if current_border else None,
            top=Side(style='thin'),
            right=current_border.right if current_border else None,
            bottom=current_border.bottom if current_border else None
        )
    
    # A～Y列の行4下部罫線
    for col_idx in range(1, 26):  # A(1)～Y(25)
        cell = ws.cell(row=4, column=col_idx)
        current_border = cell.border
        cell.border = Border(
            left=current_border.left if current_border else None,
            top=current_border.top if current_border else None,
            right=current_border.right if current_border else None,
            bottom=Side(style='thin')
        )
    
    # Y列の行3、4の右側罫線
    for row in [3, 4]:
        cell = ws.cell(row=row, column=25)
        current_border = cell.border
        cell.border = Border(
            left=current_border.left if current_border else None,
            top=current_border.top if current_border else None,
            right=Side(style='thin'),
            bottom=current_border.bottom if current_border else None
        )
    
    # 2. BN～BQ列（66～69列）の行2、3の下部に罫線追加
    for row in [2, 3]:
        for col_idx in range(66, 70):  # BN(66)～BQ(69)
            cell = ws.cell(row=row, column=col_idx)
            cell.border = Border(
                left=cell.border.left if cell.border else None,
                top=cell.border.top if cell.border else None,
                right=cell.border.right if cell.border else None,
                bottom=Side(style='thin')
            )
    
    # 2-1. BN～BQ列の行4と行5を結合（BN4:BQ5を1つに結合）
    ws.merge_cells('BN4:BQ5')
    
    # 行5の下部に罫線を追加（BN～BQ列）
    for col_idx in range(66, 70):  # BN(66)～BQ(69)
        cell = ws.cell(row=5, column=col_idx)
        current_border = cell.border
        cell.border = Border(
            left=current_border.left if current_border else None,
            top=current_border.top if current_border else None,
            right=current_border.right if current_border else None,
            bottom=Side(style='thin')
        )
    
    # 3. 行9のQ列（17）、AL列（38）の右側に罫線追加
    for col_idx in [17, 38]:
        cell = ws.cell(row=9, column=col_idx)
        cell.border = Border(
            left=cell.border.left if cell.border else None,
            top=cell.border.top if cell.border else None,
            bottom=cell.border.bottom if cell.border else None,
            right=Side(style='thin')
        )
    
    # 4. AM～BP列（38～68列）の行9～26の右側に罫線追加
    for row in range(9, 27):
        for col_idx in range(38, 69):  # AM(38)～BP(68)
            cell = ws.cell(row=row, column=col_idx)
            cell.border = Border(
                left=cell.border.left if cell.border else None,
                top=cell.border.top if cell.border else None,
                bottom=cell.border.bottom if cell.border else None,
                right=Side(style='thin')
            )
    
    # 5. 行24～26のAL列（38）の左右両側に罫線追加
    for row in range(24, 27):
        cell = ws.cell(row=row, column=38)
        cell.border = Border(
            left=Side(style='thin'),
            top=cell.border.top if cell.border else None,
            bottom=cell.border.bottom if cell.border else None,
            right=Side(style='thin')
        )
    
    # 6. 行26のAL～BQ列（38～69列）の下部に罫線追加
    for col_idx in range(38, 70):  # AL(38)～BQ(69)
        cell = ws.cell(row=26, column=col_idx)
        cell.border = Border(
            left=cell.border.left if cell.border else None,
            top=cell.border.top if cell.border else None,
            right=cell.border.right if cell.border else None,
            bottom=Side(style='thin')
        )
    
    # 6-1. G列（7列）の行24、25の右側に罫線追加
    for row in [24, 25]:
        cell = ws.cell(row=row, column=7)
        current_border = cell.border
        cell.border = Border(
            left=current_border.left if current_border else None,
            top=current_border.top if current_border else None,
            bottom=current_border.bottom if current_border else None,
            right=Side(style='thin')
        )
    
    # 7. AJ列（10列）の行27～31の右側に罫線追加
    # A27:AJ31がマージされているため、AJ列の各行に直接罫線を設定
    for row in range(27, 32):
        # AJ列 = 10列目に右側罫線を追加
        cell = ws.cell(row=row, column=10)
        # 現在の罫線を保持しながら右側罫線を追加
        current_border = cell.border
        cell.border = Border(
            left=current_border.left if current_border else None,
            top=current_border.top if current_border else None,
            bottom=current_border.bottom if current_border else None,
            right=Side(style='thin')
        )
    
    # ============================================================
    # 最終罫線調整: AJ列の行27～31の右側罫線（マージセル対応）
    # ============================================================
    
    # J列の行27～31の右側罫線を削除
    for row in range(27, 32):
        cell = ws.cell(row=row, column=10)  # J列 = 10列目
        current_border = cell.border if cell.border else Border()
        cell.border = Border(
            left=current_border.left,
            top=current_border.top,
            bottom=current_border.bottom
        )
    
    # AJ列の行27～31の右側に罫線を追加
    for row in range(27, 32):
        cell = ws.cell(row=row, column=36)  # AJ列 = 36列目
        current_border = cell.border if cell.border else Border()
        cell.border = Border(
            left=current_border.left,
            top=current_border.top,
            bottom=current_border.bottom,
            right=Side(style='thin')
        )
    
    # ============================================================
    # ファイル保存
    # ============================================================
    
    wb.save(output_path)
    print(f'✅ Excel生成成功: {output_path}')


if __name__ == '__main__':
    # コマンドライン引数からデータを取得
    if len(sys.argv) < 3:
        print('Usage: python excel_generator_advanced.py <json_data> <output_path>')
        sys.exit(1)
    
    json_data = sys.argv[1]
    output_path = sys.argv[2]
    
    # JSONデータをパース
    data = json.loads(json_data)
    
    # Excel生成
    create_inspection_report(data, output_path)

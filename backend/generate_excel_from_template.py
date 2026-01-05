#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel帳票生成スクリプト（テンプレート方式）
アップロードされたテンプレートExcelファイルを読み込み、
点検データを入力して出力します。
"""

import sys
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime
import os

def generate_excel_from_template(template_path, output_path, data):
    """
    テンプレートExcelを読み込み、点検データを入力
    
    Args:
        template_path: テンプレートExcelファイルのパス
        output_path: 出力Excelファイルのパス
        data: 点検データ（JSON）
    """
    try:
        print(f"📄 テンプレート読み込み: {template_path}")
        
        # テンプレートExcelを読み込み
        wb = load_workbook(template_path)
        ws = wb.active
        
        print(f"✅ テンプレート読み込み完了")
        print(f"📊 シート名: {ws.title}")
        
        # データを解析
        machine_type = data.get('machine_type', '')
        machine_model = data.get('machine_model', '')
        machine_unit = data.get('machine_unit', '')
        site_name = data.get('site_name', '')
        company_name = data.get('company_name', '')
        responsible_person = data.get('responsible_person', '')
        prime_contractor_inspector = data.get('prime_contractor_inspector', '')
        year = data.get('year', datetime.now().year)
        month = data.get('month', datetime.now().month)
        records = data.get('records', [])
        
        print(f"📋 データ: {machine_type} {machine_model} {machine_unit}")
        print(f"📅 対象月: {year}年{month}月")
        print(f"📝 記録数: {len(records)}件")
        
        # タイトル（A1セル）を更新
        ws['A1'] = '日々点検表'
        
        # 年月（A2セル）を更新
        ws['A2'] = f'{year}年{month}月'
        
        # 重機名（A5セル）を更新
        ws['A5'] = machine_type
        
        # 所有会社名（AM4セル）を更新
        if company_name:
            ws['AM4'] = company_name
        
        # 元請点検責任者（AX4セル）を更新
        if prime_contractor_inspector:
            ws['AX4'] = prime_contractor_inspector
        
        # 型式（BE4セル）を更新
        ws['BE4'] = machine_model
        
        # 号機（BI4セル）を更新
        ws['BI4'] = machine_unit
        
        print(f"✅ 基本情報を入力完了")
        
        # 点検結果を入力
        # AM列から開始（列インデックス39）
        # 行10～23が点検項目（行インデックス10～23）
        
        for record in records:
            day = record.get('day')
            inspector_name = record.get('inspector_name', '')
            results = record.get('results', {})
            
            if day is None:
                continue
            
            # 日付に対応する列を計算（AM列=39, AN列=40, ...）
            col_index = 38 + day  # AM列は38（0-indexed）
            
            # 列名を計算（例: AM, AN, AO, ...）
            col_name = _get_column_name(col_index)
            
            # 点検者名を入力（24～26行結合セル）
            inspector_cell = f'{col_name}24'
            if inspector_cell in ws.merged_cells:
                # 結合セルの場合、最初のセルに入力
                ws[inspector_cell] = inspector_name
            else:
                ws[inspector_cell] = inspector_name
            
            print(f"  日付{day}日: {inspector_cell} = {inspector_name}")
            
            # 点検結果を入力（10～23行）
            row_index = 10
            for item_code, result in results.items():
                if row_index > 23:
                    break
                
                cell_name = f'{col_name}{row_index}'
                
                # 結果を入力（○、×、-）
                is_good = result.get('isGood', True)
                value = '○' if is_good else '×'
                
                ws[cell_name] = value
                
                # フォント色を設定
                if is_good:
                    ws[cell_name].font = Font(color="00AA00", size=14, bold=True)
                else:
                    ws[cell_name].font = Font(color="FF0000", size=14, bold=True)
                
                # 中央揃え
                ws[cell_name].alignment = Alignment(horizontal='center', vertical='center')
                
                row_index += 1
        
        print(f"✅ 点検結果を入力完了")
        
        # ファイルを保存
        wb.save(output_path)
        print(f"💾 保存完了: {output_path}")
        
        return output_path
        
    except Exception as e:
        print(f"❌ エラー: {e}")
        import traceback
        traceback.print_exc()
        return None

def _get_column_name(col_index):
    """
    列インデックスから列名を取得（0-indexed）
    例: 0='A', 25='Z', 26='AA', 38='AM'
    """
    name = ''
    col_index += 1  # 1-indexedに変換
    while col_index > 0:
        col_index -= 1
        name = chr(col_index % 26 + ord('A')) + name
        col_index //= 26
    return name

def main():
    """
    メイン関数
    コマンドライン引数からJSONデータを受け取り、Excel生成
    """
    if len(sys.argv) < 2:
        print("❌ 使用方法: python generate_excel_from_template.py '<JSON_DATA>'")
        sys.exit(1)
    
    # JSONデータを解析
    json_data = sys.argv[1]
    data = json.loads(json_data)
    
    # テンプレートパス
    template_path = '/home/user/uploaded_files/点検表_4ｔ_1号機_2025年12月.xlsx'
    
    # 出力パス
    machine_type = data.get('machine_type', '重機')
    machine_model = data.get('machine_model', '')
    machine_unit = data.get('machine_unit', '')
    year = data.get('year', datetime.now().year)
    month = data.get('month', datetime.now().month)
    
    output_filename = f'日々点検表_{machine_type}_{machine_model}_{machine_unit}_{year}年{month}月.xlsx'
    output_path = f'/tmp/{output_filename}'
    
    # Excel生成
    result = generate_excel_from_template(template_path, output_path, data)
    
    if result:
        print(f"✅ 成功: {result}")
        # 結果をJSONで出力
        print(json.dumps({'success': True, 'file_path': result}))
    else:
        print(json.dumps({'success': False, 'error': 'Excel生成に失敗しました'}))
        sys.exit(1)

if __name__ == '__main__':
    main()
